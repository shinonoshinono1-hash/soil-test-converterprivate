from __future__ import annotations

import copy
import io
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Side

from models import Hole, Sample

COLUMNS = {
    "hole": "B", "stratum": "C", "sample_no": "D", "depth": "E",
    "wet_density": "F", "dry_density": "G", "particle_density": "H",
    "water_content": "I", "void_ratio": "J", "saturation": "K",
    "gravel": "L", "sand": "M", "silt": "N", "clay": "O", "fc": "P",
    "max_size": "Q", "d50": "R", "d20": "S", "d10": "T",
    "classification_name": "U", "classification_symbol": "V",
    "liquid_limit": "W", "plastic_limit": "X", "plasticity_index": "Y",
    "consistency_index": "Z",
    "qu1": "AA", "qu2": "AB", "qu3": "AC",
    "cu": "AD", "phi_u": "AE", "pc": "AF", "cc": "AG",
}

HEADER_MERGES = [
    "B{r}:B{r2}", "C{r}:C{r2}", "D{r}:D{r2}", "E{r}:E{r2}",
    "L{r}:T{r}", "U{r}:V{r}", "W{r}:Z{r}",
    "AA{r}:AC{r}", "AA{r2}:AC{r2}", "AD{r}:AE{r}",
]
# E（採取深度）は見本どおり2段縦結合。
# AA:AC は一軸圧縮の3入力列。AF/AG は Pc/Cc の単位行。


def build_excel(holes: list[Hole], template_path: str | Path) -> tuple[bytes, dict]:
    wb = load_workbook(template_path)
    style_wb = load_workbook(template_path)
    ws = wb["一覧表"]
    style_ws = style_wb["一覧表"]

    # 重要：
    # 先に元テンプレートの結合セルを解除・クリアしてから列追加する。
    # openpyxl は insert_cols() 後の結合セル座標を自動調整しないため、
    # 列追加を先に行うと unmerge_cells() で KeyError が発生することがある。
    _prepare_sheet(ws)

    # W:X:Y:Z（コンシステンシー）と既存AA:AB（三軸圧縮）の間に
    # 一軸圧縮用3列 AA:AC を動的に追加。既存列は右へ3列移動する。
    _insert_unconfined_columns(ws)
    _insert_unconfined_columns(style_ws)
    hole_numbers = [h.hole_no for h in holes]
    if hole_numbers:
        holes_text = "・".join(f"No.{n}" for n in hole_numbers)
        ws["B2"] = f"「高度専門教育訓練センター新築工事地質調査業務委託」室内土質試験結果一覧表（{holes_text}地点）"

    current_row = 4
    row_map: dict[tuple[int, str], int] = {}
    for hole_index, hole in enumerate(holes):
        _write_header(ws, style_ws, current_row)
        data_start = current_row + 2
        for offset, sample in enumerate(hole.samples):
            row = data_start + offset
            _copy_row_style(style_ws, 6, ws, row, 2, 33)
            _write_sample(ws, row, sample)
            _normalize_data_row_borders(ws, row)
            row_map[(hole.hole_no, sample.sample_no)] = row

        data_end = data_start + len(hole.samples) - 1
        if hole.samples:
            ws.merge_cells(start_row=data_start, start_column=2, end_row=data_end, end_column=2)
            ws.cell(data_start, 2, f"No.{hole.hole_no}")
            ws.merge_cells(start_row=data_start, start_column=3, end_row=data_end, end_column=3)
            ws.cell(data_start, 3, None)
            for col in (2, 3):
                base = copy.copy(style_ws.cell(6, col).alignment)
                ws.cell(data_start, col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=base.wrap_text,
                    text_rotation=base.text_rotation, shrink_to_fit=base.shrink_to_fit,
                )

        # 表外枠を少し強調。通常セルはテンプレートの細線を保持。
        _emphasize_block(ws, current_row, data_end)
        current_row = data_end + 1
        if hole_index < len(holes) - 1:
            # 各No.の間は完全な空白行を1行だけ。
            _clear_row(ws, current_row, 2, 33)
            current_row += 1

    # 余計な旧データは削除して、必要な高さだけにする。
    if ws.max_row > current_row + 1:
        ws.delete_rows(current_row + 1, ws.max_row - current_row)

    # 出力する数値・文字はすべて通常ウェイト（太字なし）に統一。
    _remove_all_bold(ws)
    # 最終安全処理：テンプレート由来の太字が残らないよう保存直前に再適用。
    _remove_all_bold(ws)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {"row_map": row_map, "last_row": current_row - 1}




def _insert_unconfined_columns(ws):
    """Z列の後に一軸圧縮用3列を追加し、既存表の書式を崩さずAA:ACを作る。"""
    ws.insert_cols(27, amount=3)

    # 挿入前のAA/AB（三軸圧縮）は、挿入後はAD/AEへ移動している。
    # 新AA:ACはその既存列の幅・フォント・配置・罫線・表示形式を複製し、
    # テンプレート全体の見た目を維持する。
    source_cols = {27: 30, 28: 31, 29: 30}  # AA<-AD, AB<-AE, AC<-AD
    for dst_col, src_col in source_cols.items():
        dst_letter = ws.cell(1, dst_col).column_letter
        src_letter = ws.cell(1, src_col).column_letter
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
        for r in range(1, ws.max_row + 1):
            src = ws.cell(r, src_col)
            dst = ws.cell(r, dst_col)
            dst._style = copy.copy(src._style)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
            dst.number_format = src.number_format


def _remove_all_bold(ws):
    """出力表の文字・数値をすべて確実に太字なしに統一する。"""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            font = copy.copy(cell.font)
            font.b = False
            font.bold = False
            cell.font = font


def _normalize_data_row_borders(ws, row: int):
    """各試料の横区切りはhairlineにする。No.ブロック外枠は別処理で強調する。"""
    separator = Side(style="hair", color="000000")
    # B/Cは孔番・地層記号の縦結合に使うため、データ本体 D:AD を対象にする。
    for c in range(4, 34):
        cell = ws.cell(row, c)
        b = copy.copy(cell.border)
        cell.border = Border(
            left=b.left, right=b.right,
            top=separator, bottom=separator,
            diagonal=b.diagonal, diagonal_direction=b.diagonal_direction,
            diagonalUp=b.diagonalUp, diagonalDown=b.diagonalDown,
            outline=b.outline, vertical=b.vertical, horizontal=b.horizontal,
        )

def _prepare_sheet(ws):
    # B4:AD以降にかかる既存結合をすべて解除。
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= 4 and merged.min_col <= 30 and merged.max_col >= 2:
            ws.unmerge_cells(str(merged))
    # 既存内容を削除。タイトル/列幅/印刷設定等は保持。
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=33):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_header(ws, style_ws, row: int):
    _copy_row_style(style_ws, 4, ws, row, 2, 33)
    _copy_row_style(style_ws, 5, ws, row + 1, 2, 33)
    for c in range(2, 34):
        ws.cell(row, c).value = style_ws.cell(4, c).value
        ws.cell(row + 1, c).value = style_ws.cell(5, c).value
    # 元見本では縦結合の影響で実ファイル上の2段目値が欠落する場合があるため、
    # 実務上必要な単位は明示的に復元する。
    # 採取深度の下段には線・単位を追加しない。E列はヘッダー2段を縦結合。
    ws.cell(row + 1, 5).value = None
    # Pc/Cc の下段も、既存表と同じ細字・中央揃え・罫線に統一する。
    for target_col, source_col in ((32, 30), (33, 31)):
        src = ws.cell(row + 1, source_col)
        dst = ws.cell(row + 1, target_col)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    ws.cell(row + 1, 32).value = "Pc（kN/㎡）"
    ws.cell(row + 1, 33).value = "Cc（－）"
    # 一軸圧縮：上段・下段ともAA:ACを横結合し、既存表と同じ書式に統一。
    # 見出しの書式は直前のコンシステンシー見出し（W列）を基準にする。
    for target_row, source_col in ((row, 23), (row + 1, 23)):
        src = ws.cell(target_row, source_col)
        dst = ws.cell(target_row, 27)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    ws.cell(row, 27).value = "一軸圧縮"
    ws.cell(row + 1, 27).value = "一軸圧縮強さqu (kN/m)"
    # 三軸圧縮は追加列の右側 AD:AE に移動。
    ws.cell(row, 30).value = "三軸圧縮（UU）"

    # AF/AGの単位セルは余分なスペースなし・太字なし。
    for c in (32, 33):
        font = copy.copy(ws.cell(row + 1, c).font)
        font.bold = False
        ws.cell(row + 1, c).font = font
    for fmt in HEADER_MERGES:
        ws.merge_cells(fmt.format(r=row, r2=row + 1))
    # 各No.の項目欄直下は太線に統一する。
    header_bottom = Side(style="medium", color="000000")
    for c in range(2, 34):
        cell = ws.cell(row + 1, c)
        b = copy.copy(cell.border)
        cell.border = Border(
            left=b.left, right=b.right, top=b.top, bottom=header_bottom,
            diagonal=b.diagonal, diagonal_direction=b.diagonal_direction,
            diagonalUp=b.diagonalUp, diagonalDown=b.diagonalDown,
            outline=b.outline, vertical=b.vertical, horizontal=b.horizontal,
        )
    ws.row_dimensions[row].height = max(style_ws.row_dimensions[4].height or 21, 25)
    ws.row_dimensions[row + 1].height = max(style_ws.row_dimensions[5].height or 36, 38)


def _copy_row_style(src_ws, src_row: int, dst_ws, dst_row: int, min_col: int, max_col: int):
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for c in range(min_col, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format


def _write_sample(ws, row: int, s: Sample):
    ws[f"D{row}"] = s.sample_no
    ws[f"E{row}"] = s.depth

    direct = [
        "wet_density", "dry_density", "particle_density", "water_content", "void_ratio", "saturation",
        "gravel", "sand", "max_size", "d50", "d20", "d10", "liquid_limit", "plastic_limit",
        "plasticity_index", "consistency_index", "qu1", "qu2", "qu3",
        "cu", "phi_u", "pc", "cc",
    ]
    for field in direct:
        _set_number_or_blank(ws[COLUMNS[field] + str(row)], getattr(s, field))

    ws[f"U{row}"] = s.classification_name
    ws[f"V{row}"] = s.classification_symbol

    if s.combined_fines is not None:
        ws.merge_cells(f"N{row}:O{row}")
        _set_number_or_blank(ws[f"N{row}"], s.combined_fines)
        _set_number_or_blank(ws[f"P{row}"], s.combined_fines)
    else:
        _set_number_or_blank(ws[f"N{row}"], s.silt)
        _set_number_or_blank(ws[f"O{row}"], s.clay)
        _set_number_or_blank(ws[f"P{row}"], s.fc)


def _set_number_or_blank(cell, raw: str | None):
    if raw in (None, ""):
        cell.value = None
        return
    try:
        value = Decimal(raw)
    except Exception:
        cell.value = raw
        return
    cell.value = float(value)
    decimals = len(raw.rsplit(".", 1)[1]) if "." in raw else 0
    cell.number_format = "0" if decimals == 0 else "0." + ("0" * decimals)


def _clear_row(ws, row: int, min_col: int, max_col: int):
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row, c)
        cell.value = None
        cell._style = copy.copy(ws.cell(1, 1)._style)
        cell.border = Border()
        cell.fill = copy.copy(ws.cell(1, 1).fill)


def _emphasize_block(ws, header_row: int, data_end: int):
    if data_end < header_row:
        return
    medium = Side(style="medium", color="000000")
    # 上端・下端のみ中太線。内部はテンプレートの罫線を維持。
    for c in range(2, 34):
        top_cell = ws.cell(header_row, c)
        b = copy.copy(top_cell.border)
        top_cell.border = Border(left=b.left, right=b.right, top=medium, bottom=b.bottom)
        bottom_cell = ws.cell(data_end, c)
        b2 = copy.copy(bottom_cell.border)
        bottom_cell.border = Border(left=b2.left, right=b2.right, top=b2.top, bottom=medium)
