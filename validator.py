from __future__ import annotations

from dataclasses import asdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from models import Hole, Sample
from excel_writer import COLUMNS


CHECK_FIELDS = [
    "sample_no", "depth", "wet_density", "dry_density", "particle_density", "water_content",
    "void_ratio", "saturation", "gravel", "sand", "silt", "clay", "fc", "max_size", "d50", "d20", "d10",
    "classification_name", "classification_symbol", "liquid_limit", "plastic_limit", "plasticity_index",
    "consistency_index", "cu", "phi_u", "pc", "cc",
]


def collect_warnings(holes: list[Hole]) -> list[str]:
    warnings = []
    for hole in holes:
        for s in hole.samples:
            for w in s.warnings:
                warnings.append(f"No.{hole.hole_no} {s.sample_no}: {w}")
    return warnings


def validate_generated(xlsx_bytes: bytes, holes: list[Hole], row_map: dict) -> list[dict]:
    import io
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    ws = wb["一覧表"]
    mismatches = []

    for hole in holes:
        for s in hole.samples:
            row = row_map[(hole.hole_no, s.sample_no)]
            expected = {
                "sample_no": s.sample_no,
                "depth": s.depth,
                "wet_density": s.wet_density,
                "dry_density": s.dry_density,
                "particle_density": s.particle_density,
                "water_content": s.water_content,
                "void_ratio": s.void_ratio,
                "saturation": s.saturation,
                "gravel": s.gravel,
                "sand": s.sand,
                "silt": s.silt if s.combined_fines is None else s.combined_fines,
                "clay": s.clay if s.combined_fines is None else None,
                "fc": s.fc,
                "max_size": s.max_size,
                "d50": s.d50,
                "d20": s.d20,
                "d10": s.d10,
                "classification_name": s.classification_name,
                "classification_symbol": s.classification_symbol,
                "liquid_limit": s.liquid_limit,
                "plastic_limit": s.plastic_limit,
                "plasticity_index": s.plasticity_index,
                "consistency_index": s.consistency_index,
                "cu": s.cu,
                "phi_u": s.phi_u,
                "pc": s.pc,
                "cc": s.cc,
            }
            for field, exp in expected.items():
                col = COLUMNS.get(field)
                if field == "sample_no": col = "D"
                if field == "depth": col = "E"
                actual = ws[f"{col}{row}"].value
                if not _equivalent(exp, actual):
                    mismatches.append({
                        "hole": hole.hole_no, "sample": s.sample_no, "field": field,
                        "pdf": exp, "excel": actual,
                    })
    return mismatches


def _equivalent(expected, actual) -> bool:
    if expected in (None, ""):
        return actual in (None, "")
    if isinstance(actual, str):
        if str(expected) == actual:
            return True
    try:
        return Decimal(str(expected)) == Decimal(str(actual))
    except (InvalidOperation, ValueError):
        return str(expected).strip() == str(actual).strip()
